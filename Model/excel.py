from Model.constants import *
from Model.invoices_list import InvoicesList
from openpyxl import Workbook


def create_xlsx(header: list, invoices: InvoicesList, file_name: str, xml_files: list) -> None:
    """
    Cria o arquivo XLSX com os dados extraídos da conferência.

    :param header:    Cabeçalho da tabela de dados.
    :type header:     (list)
    :param invoices:  Dados extraídos da conferência.
    :type invoices:   (InvoiceList)
    :param file_name: Nome do arquivo XLSX a ser criado.
    :type file_name:  (str)
    :param xml_files: Nomes dos arquivos XML que foram conferidos.
    :type xml_files:  (list)
    """
    # cria o arquivo Excel a ser editado
    excel_file = Workbook()

    # gera uma planilha (sheet1)
    sheet1 = excel_file.active

    sheet1.append(header)

    # inserir na planilha os dados obtidos após confirmação de lançamento do usuário
    for invoice in invoices:
        row = invoice.data_list()
        # adiciona a célula 'CANCELADA' caso essa esteja na linha e limpa os valores dela para que
        # não sejam somados ao valor total, senão copia só até a natureza
        if invoice.is_canceled:
            row = row[:header.index('Natureza') + 1]
            row.append('CANCELADA')
        else:
            row = row[:header.index('Natureza') + 1]
        # print('row:', row)
        sheet1.append(row)

    upload_sheet_content(sheet1, xml_files)

    # salva o arquivo Excel
    # excel_file.save(f'{folder.split("/")[-1]}.xlsx')
    excel_file.save(file_name)

    # abre o arquivo Excel gerado
    # os.system(f'start excel.exe {folder.split("/")[-1]}.xlsx')


def upload_sheet_content(sheet1, xml_files):
    """Edita conteúdo da planilha passada como parâmetro com a finalidade de manter a
       organização estrutural da mesma."""

    number_of_rows = len(xml_files) + 1  # +1 por conta do cabeçalho
    number_of_columns = len(HEADER1)

    # centraliza o conteudo das celulas
    from openpyxl.styles import Alignment, PatternFill
    for cell_row in range(number_of_rows + 3):  # +3 por conta dos valores de soma ao final
        for cell_column in range(len(HEADER1)):  # a primeira coluna não é centralizada
            cell = sheet1.cell(row=(cell_row + 1), column=(cell_column + 1))
            cell.alignment = Alignment(horizontal='center')

    # torna vermelha a linha de notas canceladas
    for row in range(number_of_rows):
        # column len(COLUMN_TITLES) + 1 pois vai até uma posição além do limite da coluna
        cell = sheet1.cell(row=(row + 1), column=(len(HEADER1) + 1))
        if cell.value == 'CANCELADA':
            for cell_column in range(number_of_columns):
                red = '9e1010'
                red_fill = PatternFill(start_color=red, end_color=red, fill_type='solid')
                cell = sheet1.cell(row=(row + 1), column=(cell_column + 1))
                cell.fill = red_fill
            continue

    # arruma a largura das colunas
    from openpyxl.utils import get_column_letter
    for cell_column in range(number_of_columns):
        sheet1.column_dimensions[get_column_letter(cell_column + 1)].width = 20

    # calcula a soma dos valores brutos e impostos e imprime na planilha
    gross_value_sum = 0
    iss_value_sum = 0
    ir_value_sum = 0
    csrf_value_sum = 0
    net_value_sum = 0
    gross_value_column = HEADER1.index('Vlr Bruto') + 1
    iss_value_column = HEADER1.index('Reten. ISS') + 1
    ir_value_column = HEADER1.index('Reten. IR') + 1
    csrf_value_column = HEADER1.index('Ret. CSRF') + 1
    net_value_column = HEADER1.index('Vlr Líquido') + 1
    for i in range(number_of_rows):
        gross_value_cell = sheet1.cell(row=(i + 2), column=gross_value_column).value
        iss_value_cell = sheet1.cell(row=(i + 2), column=iss_value_column).value
        ir_value_cell = sheet1.cell(row=(i + 2), column=ir_value_column).value
        csrf_value_cell = sheet1.cell(row=(i + 2), column=csrf_value_column).value
        net_value_cell = sheet1.cell(row=(i + 2), column=net_value_column).value

        if type(gross_value_cell) != str and gross_value_cell is not None:
            gross_value_sum += gross_value_cell
        if type(iss_value_cell) != str and gross_value_cell is not None:
            iss_value_sum += iss_value_cell
        if type(ir_value_cell) != str and ir_value_cell is not None:
            ir_value_sum += ir_value_cell
        if type(csrf_value_cell) != str and csrf_value_cell is not None:
            csrf_value_sum += csrf_value_cell
        if type(net_value_cell) != str and net_value_cell is not None:
            net_value_sum += net_value_cell

    sheet1.cell(row=number_of_rows + 3, column=gross_value_column).value = gross_value_sum
    sheet1.cell(row=number_of_rows + 3, column=iss_value_column).value = iss_value_sum
    sheet1.cell(row=number_of_rows + 3, column=ir_value_column).value = ir_value_sum
    sheet1.cell(row=number_of_rows + 3, column=csrf_value_column).value = csrf_value_sum
    sheet1.cell(row=number_of_rows + 3, column=net_value_column).value = net_value_sum
