from openpyxl import Workbook
from Model.constants import *
from Model.invoice import Invoice
import os


def make_excel_file() -> None:
    """Cria um arquivo Excel contendo uma planilha com os dados referentes ao servico contido na nota"""

    # cria o arquivo Excel a ser editado
    excel_file = Workbook()

    # gera uma planilha (sheet1)
    sheet1 = excel_file.active

    # insere o cabeçalho da planilha
    # 'Nº Nota', 'Data', 'Valor Bruto', 'Retenção ISS',
    # 'Retenção IR', 'Retenção CSRF', 'Valor Líquido', 'Natureza'
    sheet1.append([COLUMN_TITLES[0], COLUMN_TITLES[1], COLUMN_TITLES[2], COLUMN_TITLES[3],
                   COLUMN_TITLES[4], COLUMN_TITLES[5], COLUMN_TITLES[6], COLUMN_TITLES[7]])

    xml_files = os.listdir(XML_DIR)

    # remove os arquivos ou pastas que não forem do formato XML
    for file in xml_files:
        if '.xml' not in file:
            del(xml_files[xml_files.index(file)])

    # insere os dados de cada um dos arquivos xml a serem analisados
    for invoice_file in xml_files:
        invoice = Invoice(invoice_file)
        row = invoice.excel_data_list()
        sheet1.append(row)
    
    edit_sheet_content(sheet1, xml_files)

    # salva o arquivo Excel
    excel_file.save('output_data.xlsx')

    # abre o arquivo Excel gerado
    # os.system('start excel.exe output_data.xlsx')


def edit_sheet_content(sheet1, xml_files) -> None:
    """Edita conteúdo da planilha passada como parâmetro com a finalidade de manter a
       organização estrutural da mesma."""

    number_of_rows = len(xml_files) + 1  # +1 por conta do cabeçalho
    number_of_columns = len(COLUMN_TITLES)

    # centraliza o conteudo das celulas
    from openpyxl.styles import Alignment, PatternFill
    for cell_row in range(number_of_rows + 3):  # +3 por conta dos valores de soma ao final
        for cell_column in range(len(COLUMN_TITLES)):  # a primeira coluna não é centralizada
            cell = sheet1.cell(row=(cell_row + 1), column=(cell_column + 1))
            cell.alignment = Alignment(horizontal='center')

    # torna vermelha a linha de notas canceladas
    for cell_row in range(number_of_rows):
        cell = sheet1.cell(row=(cell_row + 1), column=(len(COLUMN_TITLES) + 1))
        if cell.value == 'CANCELADA':
            for cell_column in range(number_of_columns):
                red = '9e1010'
                red_fill = PatternFill(start_color=red, end_color=red, fill_type='solid')
                cell = sheet1.cell(row=(cell_row + 1), column=(cell_column + 1))
                cell.fill = red_fill

                # deixa os campos com os dados da nota em branco para não alterar as somas ao final
                if (cell_column + 1) > 1:
                    cell.value = None

    # arruma a largura das colunas
    from openpyxl.utils import get_column_letter
    for cell_column in range(number_of_columns):
        sheet1.column_dimensions[get_column_letter(cell_column + 1)].width = 20

    # calcula a soma dos valores brutos e impostos e imprime na planilha
    gross_value_sum = 0
    iss_value_sum = 0
    ir_value_sum = 0
    csrf_value_sum = 0
    net_value_sum = 0
    gross_value_column = COLUMN_TITLES.index('Valor Bruto') + 1
    iss_value_column = COLUMN_TITLES.index('Retenção ISS') + 1
    ir_value_column = COLUMN_TITLES.index('Retenção IR') + 1
    csrf_value_column = COLUMN_TITLES.index('Retenção CSRF') + 1
    net_value_column = COLUMN_TITLES.index('Valor Líquido') + 1
    for i in range(number_of_rows):
        gross_value_cell = sheet1.cell(row=(i + 2), column=gross_value_column).value
        iss_value_cell = sheet1.cell(row=(i + 2), column=iss_value_column).value
        ir_value_cell = sheet1.cell(row=(i + 2), column=ir_value_column).value
        csrf_value_cell = sheet1.cell(row=(i + 2), column=csrf_value_column).value
        net_value_cell = sheet1.cell(row=(i + 2), column=net_value_column).value

        if gross_value_cell:
            gross_value_sum += gross_value_cell
        if iss_value_cell:
            iss_value_sum += iss_value_cell
        if ir_value_cell:
            ir_value_sum += ir_value_cell
        if csrf_value_cell:
            csrf_value_sum += csrf_value_cell
        if net_value_cell:
            net_value_sum += net_value_cell

    sheet1.cell(row=number_of_rows + 3, column=gross_value_column).value = gross_value_sum
    sheet1.cell(row=number_of_rows + 3, column=iss_value_column).value = iss_value_sum
    sheet1.cell(row=number_of_rows + 3, column=ir_value_column).value = ir_value_sum
    sheet1.cell(row=number_of_rows + 3, column=csrf_value_column).value = csrf_value_sum
    sheet1.cell(row=number_of_rows + 3, column=net_value_column).value = net_value_sum
