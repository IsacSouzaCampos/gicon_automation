from detect_taxes_fpolis import DetectTaxes as DT
from openpyxl import Workbook
from constants import *
import os

def offline_excel_results() -> None:
    """Gera um arquivo output_data do tipo xlsx (Excel) dos resultados de análise de retenções
    das notas contidas na pasta notas_xml encontrada na pasta raiz."""

    excel_file = Workbook()
    sheet1 = excel_file.active

    sheet1.append([COLUMN1_TEXT, COLUMN2_TEXT])
    for nota in os.listdir(XML_DIR):
        if DT(nota).iss_withheld():
            sheet1.append([int(nota.split('-')[-1].replace('.xml', '')), 'X'])
        else:
            sheet1.append([int(nota.split('-')[-1].replace('.xml', '')), '-'])
    
    # centraliza o conteudo das celulas que contem informacao da retencao de ISS (coluna 2)
    from openpyxl.styles import Alignment
    for cell_row in range(len( os.listdir(XML_DIR) )):
        cell = sheet1.cell(row=(cell_row + 2), column=2)
        cell.alignment = Alignment(horizontal='center')
    
    cell = sheet1.cell(row=1, column=1)
    cell.alignment = Alignment(horizontal='center')
    
    # arruma a largura da coluna 2
    from openpyxl.utils import get_column_letter
    sheet1.column_dimensions[get_column_letter(2)].width = len(COLUMN2_TEXT)
    
    excel_file.save('output_data.xlsx')
