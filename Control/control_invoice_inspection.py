from openpyxl import Workbook
from Model.constants import *
from Model.invoice import Invoice
from View import view_invoice_inspection


def inspect_invoices(folder, xml_files) -> None:
    """Cria um arquivo Excel contendo uma planilha com os dados referentes ao servico contido na nota"""

    # cria o arquivo Excel a ser editado
    excel_file = Workbook()

    # gera uma planilha (sheet1)
    sheet1 = excel_file.active

    # insere o cabeçalho da planilha
    # 'Nº Nota', 'Data', 'Valor Bruto', 'Retenção ISS',
    # 'Retenção IR', 'Retenção CSRF', 'Valor Líquido', 'Natureza'
    sheet1.append([COLUMN_TITLES[0], COLUMN_TITLES[1], COLUMN_TITLES[2], COLUMN_TITLES[3],
                   COLUMN_TITLES[4], COLUMN_TITLES[5], COLUMN_TITLES[6], COLUMN_TITLES[7]])

    # inicia janela de carregamento
    window = view_invoice_inspection.start_loading_inspection_window()

    # insere os dados de cada um dos arquivos xml a serem analisados
    import time
    for i in range(len(xml_files)):
        invoice_file = xml_files[i]
        invoice = Invoice(folder, invoice_file)
        view_invoice_inspection.update_loading_window(window, invoice.d['numeroserie'], i, len(xml_files))
        row = invoice.excel_data_list()
        sheet1.append(row)
        time.sleep(.5)
    
    edit_sheet_content(sheet1, xml_files)

    # salva o arquivo Excel
    excel_file.save('output_data.xlsx')

    # abre o arquivo Excel gerado
    # os.system('start excel.exe output_data.xlsx')


def edit_sheet_content(sheet1, xml_files) -> None:
    """Edita conteúdo da planilha passada como parâmetro com a finalidade de manter a
       organização estrutural da mesma."""

    number_of_rows = len(xml_files) + 1  # +1 por conta do cabeçalho
    number_of_columns = len(COLUMN_TITLES)

    # centraliza o conteudo das celulas
    from openpyxl.styles import Alignment, PatternFill
    for cell_row in range(number_of_rows + 3):  # +3 por conta dos valores de soma ao final
        for cell_column in range(len(COLUMN_TITLES)):  # a primeira coluna não é centralizada
            cell = sheet1.cell(row=(cell_row + 1), column=(cell_column + 1))
            cell.alignment = Alignment(horizontal='center')

    # torna vermelha a linha de notas canceladas
    for cell_row in range(number_of_rows):
        cell = sheet1.cell(row=(cell_row + 1), column=(len(COLUMN_TITLES) + 1))
        if cell.value == 'CANCELADA':
            for cell_column in range(number_of_columns):
                red = '9e1010'
                red_fill = PatternFill(start_color=red, end_color=red, fill_type='solid')
                cell = sheet1.cell(row=(cell_row + 1), column=(cell_column + 1))
                cell.fill = red_fill

                # deixa os campos com os dados da nota em branco para não alterar as somas ao final
                if (cell_column + 1) > 1:
                    cell.value = None

    # arruma a largura das colunas
    from openpyxl.utils import get_column_letter
    for cell_column in range(number_of_columns):
        sheet1.column_dimensions[get_column_letter(cell_column + 1)].width = 20

    # calcula a soma dos valores brutos e impostos e imprime na planilha
    gross_value_sum = 0
    iss_value_sum = 0
    ir_value_sum = 0
    csrf_value_sum = 0
    net_value_sum = 0
    gross_value_column = COLUMN_TITLES.index('Valor Bruto') + 1
    iss_value_column = COLUMN_TITLES.index('Retenção ISS') + 1
    ir_value_column = COLUMN_TITLES.index('Retenção IR') + 1
    csrf_value_column = COLUMN_TITLES.index('Retenção CSRF') + 1
    net_value_column = COLUMN_TITLES.index('Valor Líquido') + 1
    for i in range(number_of_rows):
        gross_value_cell = sheet1.cell(row=(i + 2), column=gross_value_column).value
        iss_value_cell = sheet1.cell(row=(i + 2), column=iss_value_column).value
        ir_value_cell = sheet1.cell(row=(i + 2), column=ir_value_column).value
        csrf_value_cell = sheet1.cell(row=(i + 2), column=csrf_value_column).value
        net_value_cell = sheet1.cell(row=(i + 2), column=net_value_column).value

        if type(gross_value_cell) != str and gross_value_cell is not None:
            gross_value_sum += gross_value_cell
        if type(iss_value_cell) != str and gross_value_cell is not None:
            iss_value_sum += iss_value_cell
        if type(ir_value_cell) != str and ir_value_cell is not None:
            ir_value_sum += ir_value_cell
        if type(csrf_value_cell) != str and csrf_value_cell is not None:
            csrf_value_sum += csrf_value_cell
        if type(net_value_cell) != str and net_value_cell is not None:
            net_value_sum += net_value_cell

    sheet1.cell(row=number_of_rows + 3, column=gross_value_column).value = gross_value_sum
    sheet1.cell(row=number_of_rows + 3, column=iss_value_column).value = iss_value_sum
    sheet1.cell(row=number_of_rows + 3, column=ir_value_column).value = ir_value_sum
    sheet1.cell(row=number_of_rows + 3, column=csrf_value_column).value = csrf_value_sum
    sheet1.cell(row=number_of_rows + 3, column=net_value_column).value = net_value_sum
