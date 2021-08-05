from openpyxl import Workbook
from constants import *
from invoice import Invoice
import os


def main():
    # cria o arquivo Excel a ser editado
    excel_file = Workbook()

    # gera uma planilha (sheet1)
    sheet1 = excel_file.active

    # insere o cabecalho da planilha
    # 'Nota', 'Razão Social Tomador', 'Razão Social Prestador', 'Renteção de ISS', 'Retenção IR', 'Retenção CSRF'
    sheet1.append([COLUMN_TITLES[0], COLUMN_TITLES[1], COLUMN_TITLES[2], COLUMN_TITLES[3], COLUMN_TITLES[4], COLUMN_TITLES[5]])

    # insere os dados de cada um dos arquivos xml a serem analisados
    for invoice_file in os.listdir(XML_DIR):
        if '.xml' not in invoice_file:
            continue
        invoice = Invoice(invoice_file)
        row = invoice.excel_data_list()
        sheet1.append(row)
    
    # centraliza o conteudo das celulas das colunas 4 até a última
    from openpyxl.styles import Alignment
    for cell_row in range(len( os.listdir(XML_DIR) )):  # -1 temporário (até remover pasta com_retencao_federal)
        for cell_column in range(len(COLUMN_TITLES) - 3):  # as 3 primeiras colunas não são centralizadas
            cell = sheet1.cell(row=(cell_row + 2), column=((cell_column + 1) + 3))
            cell.alignment = Alignment(horizontal='center')
    
    # arruma a coluna 1 (largura e alinhamento)
    sheet1.column_dimensions['A'].width = 10
    for cell_row in range(len( os.listdir(XML_DIR) ) - 1):  # -1 temporário (até remover pasta com_retencao_federal)
        cell = sheet1.cell(row=(cell_row + 2), column=1)
        cell.alignment = Alignment(horizontal='left')
    
    # centraliza linha 1
    for cell_column in range(len(COLUMN_TITLES)):
        cell = sheet1.cell(row=1, column=(cell_column + 1))
        cell.alignment = Alignment(horizontal='center')
    
    # arruma a largura das colunas 2 até a última
    from openpyxl.utils import get_column_letter
    for cell_column in range(len(COLUMN_TITLES) - 1):
        sheet1.column_dimensions[get_column_letter(cell_column + 2)].width = len(COLUMN_TITLES[cell_column + 1]) + 2

    # salva o arquivo Excel
    excel_file.save('output_data.xlsx')


if __name__ == '__main__':
    main()
