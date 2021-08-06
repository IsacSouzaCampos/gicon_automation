from openpyxl import Workbook
from constants import *
from invoice import Invoice
import os


def make_excel_file() -> None:
    """Cria um arquivo Excel contendo uma planilha com os dados: 'Nº Nota', 'Data',
       'Valor Bruto', 'Renteção de ISS', 'Retenção IR', 'Retenção CSRF', 'Valor Líquido'"""

    # cria o arquivo Excel a ser editado
    excel_file = Workbook()

    # gera uma planilha (sheet1)
    sheet1 = excel_file.active

    # insere o cabecalho da planilha
    # 'Nº Nota', 'Data', 'Valor Bruto', 'Renteção de ISS',
    # 'Retenção IR', 'Retenção CSRF'
    sheet1.append([COLUMN_TITLES[0], COLUMN_TITLES[1], COLUMN_TITLES[2], COLUMN_TITLES[3],
                   COLUMN_TITLES[4], COLUMN_TITLES[5]])

    # insere os dados de cada um dos arquivos xml a serem analisados
    for invoice_file in os.listdir(XML_DIR):
        if '.xml' not in invoice_file:
            continue
        invoice = Invoice(invoice_file)
        row = invoice.excel_data_list()
        sheet1.append(row)
    
    edit_sheet_content(sheet1)

    # salva o arquivo Excel
    excel_file.save('output_data.xlsx')

    # abre o arquivo Excel gerado
    # os.system(f'start excel.exe output_data.xlsx')


def edit_sheet_content(sheet1) -> None:
    """Edita conteúdo da planilha passada como parâmetro com a finalidade de manter a
       organização estrutural da mesma."""

     # centraliza o conteudo das celulas
    from openpyxl.styles import Alignment, Color, PatternFill
    for cell_row in range(len( os.listdir(XML_DIR) ) + 1):  # +1 por conta do cabeçalho
        for cell_column in range(len(COLUMN_TITLES)):  # a primeira coluna não é centralizada
            cell = sheet1.cell(row=(cell_row + 1), column=((cell_column + 1)))
            cell.alignment = Alignment(horizontal='center')

    # torna vermelha a linha de notas canceladas
    for cell_row in range(len( os.listdir(XML_DIR) ) + 1):  # +1 por conta do cabeçalho
        cell = sheet1.cell(row=(cell_row + 1), column=(len(COLUMN_TITLES) + 1))
        # print(cell.value, cell_row + 1, len(COLUMN_TITLES) + 1)
        if cell.value == 'CANCELADA':
            for cell_column in range(len(COLUMN_TITLES)):
                red = '9e1010'
                red_fill = PatternFill(start_color=red, end_color=red, fill_type='solid')
                cell = sheet1.cell(row=(cell_row + 1), column=(cell_column + 1))
                cell.fill = red_fill

    # arruma a largura das colunas
    from openpyxl.utils import get_column_letter
    for cell_column in range(len(COLUMN_TITLES)):
        sheet1.column_dimensions[get_column_letter(cell_column + 1)].width = 20
